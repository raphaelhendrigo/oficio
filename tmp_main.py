import os
import re
import sys
import time
import unicodedata
from pathlib import Path
from urllib.parse import urljoin

from dotenv import load_dotenv
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError


def env_bool(name: str, default: bool = False) -> bool:
    v = os.getenv(name, str(default))
    return str(v).strip().lower() in ("1", "true", "yes", "y", "on")


def find_frame_with_text(page, text: str, timeout_ms: int = 30000):
    """Loop through frames until one contains the given text (substring)."""
    deadline = time.time() + (timeout_ms / 1000.0)
    last_err = None
    while time.time() < deadline:
        for fr in page.frames:
            try:
                loc = fr.get_by_text(text, exact=False)
                # .count() waits for DOM stability enough for text lookup
                if loc.count() > 0:
                    return fr
            except Exception as e:
                last_err = e
                continue
        time.sleep(0.3)
    if last_err:
        raise last_err
    raise PWTimeoutError(f"Frame with text '{text}' not found in {timeout_ms}ms.")


def find_frame_with_selector(page, selector: str, timeout_ms: int = 30000):
    """Find a frame containing an element matching selector that is attached in DOM."""
    deadline = time.time() + (timeout_ms / 1000.0)
    while time.time() < deadline:
        for fr in page.frames:
            try:
                loc = fr.locator(selector)
                if loc.count() > 0:
                    try:
                        loc.first.wait_for(state="attached", timeout=1000)
                    except Exception:
                        pass
                    return fr
            except Exception:
                continue
        time.sleep(0.3)
    raise PWTimeoutError(f"Frame with selector '{selector}' not found in {timeout_ms}ms.")


def normalize(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace("º", "o").replace("°", "o").replace("ª", "a")
    s = unicodedata.normalize("NFKD", s)
    s = "".join([c for c in s if not unicodedata.combining(c)])
    return s


def find_latest_export_file(directory: Path) -> Path | None:
    candidates = []
    for ext in ("*.xlsx", "*.xls"):
        candidates.extend(directory.glob(ext))
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def find_processo_column_index(headers: list[str]) -> int | None:
    best_idx = None
    for i, h in enumerate(headers):
        hl = normalize(h).strip().lower()
        if not hl:
            continue
        if "processo" in hl and any(tag in hl for tag in ("n", "no", "n.", "n ", "numero")):
            return i
        if best_idx is None and "processo" in hl:
            best_idx = i
    if best_idx is not None:
        return best_idx
    if len(headers) >= 5:
        return 4
    return None


def extract_processo_from_excel(path: Path) -> str | None:
    suffix = path.suffix.lower()
    try:
        if suffix == ".xlsx":
            from openpyxl import load_workbook  # type: ignore
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            ws = wb.active
            header = None
            idx = None
            for row in ws.iter_rows(values_only=True):
                values = ["" if v is None else str(v) for v in row]
                if header is None:
                    header = values
                    idx = find_processo_column_index(header)
                    if idx is None:
                        continue
                    continue
                if idx is None or idx >= len(values):
                    continue
                val = values[idx]
                if val and str(val).strip():
                    return str(val).strip()
            return None
        elif suffix == ".xls":
            import xlrd  # type: ignore
            book = xlrd.open_workbook(str(path))
            sheet = book.sheet_by_index(0)
            header_row = 0
            idx = None
            for r in range(min(5, sheet.nrows)):
                row_vals = [str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
                idx_try = find_processo_column_index(row_vals)
                if idx_try is not None:
                    header_row = r
                    idx = idx_try
                    break
            if idx is None:
                return None
            for r in range(header_row + 1, sheet.nrows):
                try:
                    v = sheet.cell_value(r, idx)
                except Exception:
                    continue
                if v is None:
                    continue
                txt = str(v).strip()
                if txt:
                    return txt
            return None
        else:
            return None
    except Exception as e:
        print(f"Aviso: falha ao ler planilha {path.name}: {e}")
        return None


def search_processo_and_open_viewer(context, page, processo: str):
    search_frame = None
    try:
        search_frame = find_frame_with_selector(page, "#cbbProcesso_I", timeout_ms=15000)
    except Exception:
        if page.locator("#cbbProcesso_I").count() == 0:
            raise

    target = search_frame if search_frame else page
    target.locator("#cbbProcesso_I").fill(processo)
    clicked = False
    pages_before = list(context.pages)
    try:
        btn = target.locator("button[onclick='BuscaProcesso();']").first
        if btn.count() > 0:
            btn.click()
            clicked = True
    except Exception:
        pass
    if not clicked:
        try:
            target.locator("#cbbProcesso_I").press("Enter")
            clicked = True
        except Exception:
            pass
    # If viewer loads in same page, its frame should appear
    try:
        find_frame_with_selector(page, "#splLeitorDocumentos_pgcPecas_trePecas", timeout_ms=60000)
        return page
    except Exception:
        pass

    # Otherwise try to detect a newly opened page
    deadline = time.time() + 10
    while time.time() < deadline:
        pages_now = list(context.pages)
        if len(pages_now) > len(pages_before):
            newp = pages_now[-1]
            try:
                newp.wait_for_load_state("domcontentloaded", timeout=5000)
            except Exception:
                pass
            return newp
        time.sleep(0.3)
    return page


def click_last_piece_and_open_pdf(context, page, output_dir: Path, processo: str) -> Path | None:
    try:
        viewer_frame = find_frame_with_selector(page, "#splLeitorDocumentos_pgcPecas_trePecas", timeout_ms=60000)
    except Exception:
        print("Aviso: Visualizador de documentos nao encontrado.")
        return None

    loc = viewer_frame.locator("a[index_ato]")
    count = loc.count()
    if count == 0:
        loc = viewer_frame.locator("a[cod_arquivo_digital_criptografado]")
        count = loc.count()
    if count == 0:
        print("Aviso: Nenhuma peca encontrada no visualizador.")
        return None

    max_idx = -1
    max_n = 0
    for i in range(count):
        item = loc.nth(i)
        val = item.get_attribute("index_ato") or item.get_attribute("index") or ""
        try:
            iv = int(re.findall(r"\d+", val)[0]) if val else i
        except Exception:
            iv = i
        if iv >= max_idx:
            max_idx = iv
            max_n = i

    loc.nth(max_n).click()
    time.sleep(1.0)

    btn_sel = "#imgNewWindow, img#imgNewWindow"
    btn_clicked = False
    try:
        if viewer_frame.locator(btn_sel).count() > 0:
            with page.expect_popup(timeout=15000) as pop_info:
                viewer_frame.locator(btn_sel).first.click()
            pdf_page = pop_info.value
            btn_clicked = True
        else:
            raise Exception("not in frame")
    except Exception:
        if page.locator(btn_sel).count() > 0:
            with page.expect_popup(timeout=15000) as pop_info:
                page.locator(btn_sel).first.click()
            pdf_page = pop_info.value
            btn_clicked = True

    if not btn_clicked:
        print("Aviso: Botao 'abrir em nova janela' nao encontrado.")
        return None

    pdf_page.wait_for_load_state("domcontentloaded", timeout=30000)

    def _pick_attr(pl, sel, attr):
        el = pl.locator(sel)
        if el.count() > 0:
            val = el.first.get_attribute(attr)
            if val:
                return val
        return None

    pdf_url = (
        _pick_attr(pdf_page, "iframe[src*='.pdf']", "src")
        or _pick_attr(pdf_page, "embed[type*='pdf']", "src")
        or _pick_attr(pdf_page, "object[data*='.pdf']", "data")
        or _pick_attr(pdf_page, "a[href*='.pdf']", "href")
    )
    if not pdf_url:
        pdf_url = _pick_attr(pdf_page, "iframe", "src")

    if not pdf_url:
        print("Aviso: URL do PDF nao encontrada.")
        return None

    abs_url = urljoin(pdf_page.url, pdf_url)
    try:
        resp = context.request.get(abs_url, timeout=60000)
        if resp.ok:
            pdf_path = output_dir / f"{processo}-ultimo-ato.pdf"
            with open(pdf_path, "wb") as f:
                f.write(resp.body())
            print(f"PDF salvo em: {pdf_path.resolve()}")
            return pdf_path
        else:
            print(f"Aviso: falha ao baixar PDF (status {resp.status}).")
            return None
    except Exception as e:
        print(f"Aviso: erro ao baixar PDF: {e}")
        return None


def main():
    load_dotenv()  # load .env if present

    url = os.getenv("ETCM_URL", "https://homologacao-etcm.tcm.sp.gov.br/paginas/login.aspx")
    username = os.getenv("ETCM_USERNAME")
    password = os.getenv("ETCM_PASSWORD")
    headless = env_bool("HEADLESS", False)

    if not username or not password:
        print("ERRO: defina ETCM_USERNAME e ETCM_PASSWORD (via .env ou variaveis de ambiente).")
        sys.exit(2)

    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless, channel="chrome")
        context = browser.new_context(viewport={"width": 1600, "height": 900}, accept_downloads=True)
        page = context.new_page()

        # 1) Login
        print(f"Acessando: {url}")
        page.goto(url, wait_until="load", timeout=60000)

        # Username: prefer explicit DevExpress ids/names; fallback strategies next
        try:
            page.locator("#ctl00_cphMain_txtUsuario_I, input[name='ctl00$cphMain$txtUsuario']").first.fill(username)
        except Exception:
            try:
                page.locator("input[placeholder*='Usu'] , input[name*='Usuario' i]").first.fill(username)
            except Exception:
                page.locator("input[type='text']").first.fill(username)

        # Password: prefer explicit DevExpress ids/names; fallback strategies next
        try:
            page.locator("#ctl00_cphMain_txtSenha_I, input[name='ctl00$cphMain$txtSenha'][type='password']").first.fill(password)
        except Exception:
            try:
                page.locator("input[type='password']").first.fill(password)
            except Exception as e:
                raise RuntimeError("Nao foi possivel localizar o campo de senha.") from e

        # Login button (DevExpress): prefer container, then other fallbacks, then Enter key
        clicked = False
        try:
            page.locator("#ctl00_cphMain_btnLogin").click()
            clicked = True
        except Exception:
            pass
        if not clicked:
            try:
                page.get_by_role("button", name=re.compile(r"Entrar|Acessar|Login", re.I)).click()
                clicked = True
            except Exception:
                pass
        if not clicked:
            try:
                page.locator(
                    "input[type='submit'][value*='Entrar' i]:not([readonly]):not([disabled])"
                ).first.click()
                clicked = True
            except Exception:
                pass
        if not clicked:
            try:
                page.locator("#ctl00_cphMain_btnLogin span.dx-vam:has-text('Entrar')").first.click()
                clicked = True
            except Exception:
                try:
                    page.locator("text=/\\b(Entrar|Acessar|Login)\\b/i").locator(":visible").first.click()
                    clicked = True
                except Exception:
                    pass
        if not clicked:
            try:
                page.locator("#ctl00_cphMain_txtSenha_I, input[type='password']").first.press("Enter")
                clicked = True
            except Exception as e:
                raise RuntimeError("Nao foi possivel acionar o login (botao/Enter nao disponiveis).") from e

        # Wait for the next page or app area to load
        page.wait_for_load_state("networkidle", timeout=60000)

        # 2) Find the menu frame with "Processos"
        print("Procurando o menu 'Processos'...")
        fr_menu = find_frame_with_text(page, "Processos", timeout_ms=30000)

        # 3) Click on "Processos"
        print("Clicando em 'Processos'...")
        fr_menu.get_by_text("Processos", exact=True).first.click(force=True)

        # 4) Click on "Em confeccao APO-PEN" (text may include counters)
        print("Abrindo 'Em confeccao APO-PEN'...")
        fr_menu.get_by_text(re.compile(r"Em\s*confec.*APO-?PEN", re.I)).first.click(force=True)

        # 5) After opening, a grid usually appears elsewhere; wait and try to export
        time.sleep(1.0)
        target_frame = None
        try:
            target_frame = find_frame_with_text(page, "Aposentadoria", timeout_ms=10000)
        except Exception:
            try:
                target_frame = find_frame_with_text(page, "Processos", timeout_ms=5000)
            except Exception:
                pass

        # 6) Click "Exportar" and download the spreadsheet (Processos grid preferred)
        print("Procurando botão 'Exportar'...")
        def _visible_count(sel: str) -> int:
            try:
                return page.locator(sel).filter(has_text="Exportar").and_(page.locator(":visible")).count()
            except Exception:
                return 0

        # Prefer the Processos grid export, then fallback to Documentos
        export_selectors = [
            "#sptMesaTrabalho_gvProcesso_Title_btnExport, #sptMesaTrabalho_gvProcesso_Title_btnExport_I",
            "#sptMesaTrabalho_gvDocumentos_Title_btnExport, #sptMesaTrabalho_gvDocumentos_Title_btnExport_I",
        ]
        export_clicked = False
        downloaded_file: Path | None = None
        for sel in export_selectors:
            try:
                loc = page.locator(sel).first
                # Wait a bit for visibility
                loc.wait_for(state="visible", timeout=15000)
                with page.expect_download(timeout=60000) as dl_info:
                    loc.click()
                download = dl_info.value
                suggested = download.suggested_filename
                dest_path = output_dir / suggested
                try:
                    download.save_as(str(dest_path))
                except Exception:
                    # Fallback to temporary path if save_as fails
                    tmp_path = download.path()
                    if tmp_path:
                        # Copy to dest
                        import shutil
                        shutil.copyfile(tmp_path, dest_path)
                print(f"Planilha baixada em: {dest_path.resolve()}")
                export_clicked = True
                downloaded_file = dest_path
                break
            except Exception:
                continue
        if not export_clicked:
            print("Aviso: botão 'Exportar' não encontrado ou download não disparou.")

        # Screenshot for sanity
        png_path = output_dir / "apos_apo_pen.png"
        page.screenshot(path=str(png_path), full_page=True)
        print(f"Screenshot salvo em: {png_path.resolve()}")

        # 7) Ler planilha e tentar baixar PDF do ultimo ato do processo
        processo_num = None
        try:
            src_file = downloaded_file if downloaded_file and downloaded_file.exists() else find_latest_export_file(output_dir)
        except Exception:
            src_file = None
        if src_file and src_file.exists():
            processo_num = extract_processo_from_excel(src_file)
            if processo_num:
                print(f"Processo identificado na planilha: {processo_num}")
            else:
                print("Aviso: nao foi possivel extrair 'N° Processo' da planilha.")
        else:
            print("Aviso: nenhuma planilha encontrada para leitura.")

        if processo_num:
            try:
                active_page = search_processo_and_open_viewer(context, page, processo_num)
                click_last_piece_and_open_pdf(context, active_page, output_dir, processo_num)
            except Exception as e:
                print(f"Aviso: falha ao navegar e baixar PDF: {e}")

        print("Concluido com sucesso.")
        context.close()
        browser.close()


if __name__ == "__main__":
    main()

