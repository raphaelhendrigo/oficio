import os
import re
import time
import unicodedata
from pathlib import Path
from typing import Any, Optional

import yaml
from playwright.sync_api import sync_playwright

from config import load_config
from logger import init_logger
from selectors import DEVEXPRESS_LOADING_SELECTORS


def slugify(text: str) -> str:
    text = re.sub(r"[^\w\-\.]+", "_", text.strip())
    text = re.sub(r"_+", "_", text).strip("_")
    return text or "step"


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def read_yaml_steps(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise FileNotFoundError(f"Steps file not found: {path}")
    data = yaml.safe_load(path.read_text(encoding="utf-8"))
    if not data:
        return []
    if not isinstance(data, list):
        raise ValueError("steps.yaml must be a list of steps")
    return data


def parse_kv_pairs(text: str) -> dict[str, str]:
    pairs: dict[str, str] = {}
    for part in text.split(";"):
        part = part.strip()
        if not part or "=" not in part:
            continue
        key, value = part.split("=", 1)
        pairs[key.strip()] = value.strip()
    return pairs


def split_candidates(hint: str) -> list[str]:
    return [h.strip() for h in hint.split("|") if h.strip()]


def expand_value(value: Any, variables: dict[str, str]) -> Any:
    if not isinstance(value, str):
        return value

    def repl(match: re.Match[str]) -> str:
        key = match.group(1)
        if key in variables:
            return str(variables[key])
        return os.getenv(key, "")

    return re.sub(r"\$\{([A-Za-z0-9_]+)\}", repl, value)


def find_latest_excel(download_dir: Path) -> Optional[Path]:
    candidates = list(download_dir.glob("*.xlsx")) + list(download_dir.glob("*.xls"))
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def extract_processes_from_excel(path: Path) -> list[str]:
    suffix = path.suffix.lower()
    processos: list[str] = []
    if suffix == ".xlsx":
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        ws = wb.active
        header = None
        idx = None
        for row in ws.iter_rows(values_only=True):
            values = ["" if v is None else str(v) for v in row]
            if header is None:
                header = values
                idx = find_processo_column_index(header)
                continue
            if idx is None or idx >= len(values):
                continue
            val = values[idx]
            if val and str(val).strip():
                processos.append(str(val).strip())
    elif suffix == ".xls":
        import xlrd  # type: ignore

        book = xlrd.open_workbook(str(path))
        sheet = book.sheet_by_index(0)
        header_row = 0
        idx = None
        for r in range(min(5, sheet.nrows)):
            row_vals = [str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            idx_try = find_processo_column_index(row_vals)
            if idx_try is not None:
                header_row = r
                idx = idx_try
                break
        if idx is None:
            return []
        for r in range(header_row + 1, sheet.nrows):
            try:
                v = sheet.cell_value(r, idx)
            except Exception:
                continue
            if v is None:
                continue
            txt = str(v).strip()
            if txt:
                processos.append(txt)
    return processos


def find_processo_column_index(headers: list[str]) -> Optional[int]:
    best_idx = None
    for i, h in enumerate(headers):
        hl = normalize(h).strip().lower()
        if not hl:
            continue
        if "processo" in hl and any(tag in hl for tag in ("n", "no", "n.", "n ", "numero")):
            return i
        if best_idx is None and "processo" in hl:
            best_idx = i
    if best_idx is not None:
        return best_idx
    return None


def normalize(text: str) -> str:
    if text is None:
        return ""
    text = str(text)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("§", "o").replace("ø", "o").replace("¦", "a")
    text = re.sub(r"\s+", " ", text)
    return text


class StepRunner:
    def __init__(self, page, context, logger, config):
        self.page = page
        self.context = context
        self.logger = logger
        self.config = config
        self.page_stack: list[Any] = []
        self.variables = {
            "BASE_URL": config.base_url,
            "ETCM_USER": config.etcm_user,
            "ETCM_PASS": config.etcm_pass,
            "DESTINATARIO": config.destinatario,
            "RELATOR": config.relator,
            "DESCRICAO": config.descricao,
            "STATUS_ENTREGA": config.status_entrega,
            "PRAZO_DIAS": config.prazo_dias,
            "ANEXO_DOCX": config.anexo_docx,
            "VISOES": config.visoes,
            "DISTRIBUIDO_PARA": config.distribuido_para,
        }

    def set_process(self, processo: str) -> None:
        self.variables["PROCESSO"] = processo

    def wait_for_idle(self, timeout_ms: int) -> None:
        selectors = ", ".join(DEVEXPRESS_LOADING_SELECTORS)
        script = (
            "sel => { const el = document.querySelector(sel);"
            " if (!el) return true;"
            " const style = window.getComputedStyle(el);"
            " return style && style.display === 'none'; }"
        )
        try:
            self.page.wait_for_function(script, arg=selectors, timeout=timeout_ms)
        except Exception:
            pass
        try:
            self.page.get_by_text("Carregando", exact=False).first.wait_for(state="hidden", timeout=timeout_ms)
        except Exception:
            pass

    def find_input_by_label(self, label_text: str):
        label = self.page.locator("label", has_text=re.compile(label_text, re.I)).first
        if label.count() == 0:
            return None
        for_attr = label.get_attribute("for")
        if for_attr:
            return self.page.locator(f"#{for_attr}")
        parent = label.locator("xpath=..")
        candidate = parent.locator("input, textarea, select").first
        if candidate.count() > 0:
            return candidate
        return None

    def find_grid_filter_input(self, grid_selector: str, column_text: str):
        grid = self.page.locator(grid_selector).first
        if grid.count() == 0:
            return None
        header_row = grid.locator("tr[id*='DXHeadersRow'], tr.dxgvHeader").first
        if header_row.count() == 0:
            return None
        headers = header_row.locator("th, td")
        try:
            header_texts = headers.all_text_contents()
        except Exception:
            header_texts = []
        target_norm = normalize(column_text).strip().lower()
        idx = None
        for i, text in enumerate(header_texts):
            if target_norm and target_norm in normalize(text).strip().lower():
                idx = i
                break
        if idx is None:
            return None
        filter_row = grid.locator("tr[id*='DXFilterRow'], tr.dxgvFilterRow").first
        if filter_row.count() == 0:
            return None
        cell = filter_row.locator("td").nth(idx)
        return cell.locator("input, textarea").first

    def resolve_locator(self, strategy: str, hint: str):
        if not hint:
            return None
        candidates = split_candidates(hint)
        last = None
        for cand in candidates:
            if strategy == "css":
                loc = self.page.locator(cand)
            elif strategy == "text":
                if cand.startswith("re:"):
                    loc = self.page.get_by_text(re.compile(cand[3:], re.I))
                else:
                    loc = self.page.get_by_text(cand, exact=False)
            elif strategy == "role":
                parts = parse_kv_pairs(cand)
                role = parts.get("role", "button")
                name = parts.get("name", "")
                name_re = re.compile(name, re.I) if name else None
                loc = self.page.get_by_role(role, name=name_re)
            elif strategy == "id_or_label":
                if cand.lower().startswith("label:"):
                    label_text = cand.split(":", 1)[1].strip()
                    loc = self.find_input_by_label(label_text)
                    if not loc:
                        continue
                else:
                    loc = self.page.locator(cand)
            elif strategy == "row":
                kv = parse_kv_pairs(cand)
                row_selector = kv.get("row_selector", "tr")
                cell_text = kv.get("cell_text", "")
                cell_text = expand_value(cell_text, self.variables)
                loc = self.page.locator(row_selector).filter(has_text=re.compile(cell_text, re.I)).first
            else:
                loc = self.page.locator(cand)
            last = loc
            try:
                if loc.count() > 0:
                    return loc
            except Exception:
                continue
        return last

    def resolve_row_icon(self, hint: str):
        kv = parse_kv_pairs(hint)
        row_selector = kv.get("row_selector", "tr")
        cell_text = kv.get("cell_text", "")
        cell_text = expand_value(cell_text, self.variables)
        icon_css = kv.get("icon_css", "")
        row = self.page.locator(row_selector).filter(has_text=re.compile(cell_text, re.I)).first
        if not icon_css:
            return row
        return row.locator(icon_css).first

    def take_evidence(self, step_id: Any, label: str) -> None:
        ensure_dir(self.config.evidence_dir)
        slug = slugify(label)
        path = self.config.evidence_dir / f"step_{step_id}_{slug}.png"
        try:
            self.page.screenshot(path=str(path), full_page=True)
            self.logger.info("Evidence saved: %s", path)
        except Exception as exc:
            self.logger.warning("Failed to capture evidence: %s", exc)

    def save_failure_artifacts(self, step_id: Any, label: str) -> None:
        ensure_dir(self.config.evidence_dir)
        ensure_dir(self.config.html_dir)
        slug = slugify(label)
        shot_path = self.config.evidence_dir / f"step_{step_id}_{slug}_error.png"
        html_path = self.config.html_dir / f"step_{step_id}_{slug}_error.html"
        try:
            self.page.screenshot(path=str(shot_path), full_page=True)
        except Exception:
            pass
        try:
            html_path.write_text(self.page.content(), encoding="utf-8")
        except Exception:
            pass
        self.logger.info("Failure artifacts: %s, %s", shot_path, html_path)

    def run_step(self, step: dict[str, Any], mode: str) -> None:
        step_id = step.get("step_id", "")
        action = step.get("action", "").strip().lower()
        target_hint = step.get("target_hint", "")
        locator_strategy = step.get("locator_strategy", "css")
        locator_hint = step.get("locator_hint", "")
        optional = bool(step.get("optional", False))
        evidence = bool(step.get("evidence", False))
        retries = int(step.get("retries", 2))
        expect_popup = bool(step.get("expect_popup", False))
        wait_state = step.get("wait_state") or step.get("state") or "visible"
        skip_if_empty = bool(step.get("skip_if_empty", False))

        value = step.get("value", "")
        value = expand_value(value, self.variables)
        if skip_if_empty and isinstance(value, str) and not value.strip():
            self.logger.info("Step %s skipped (empty value)", step_id)
            return
        locator = None
        if locator_hint:
            locator = self.resolve_locator(locator_strategy, expand_value(locator_hint, self.variables))

        for attempt in range(1, retries + 1):
            try:
                self.logger.info("Step %s | %s | %s", step_id, action, target_hint)

                if action == "goto":
                    url = value or self.config.base_url
                    self.page.goto(url, wait_until="domcontentloaded", timeout=self.config.timeout_ms)
                elif action == "click":
                    if not locator:
                        raise RuntimeError("Locator not resolved")
                    if mode != "dry-run":
                        if expect_popup:
                            with self.context.expect_page(timeout=self.config.timeout_ms) as pinfo:
                                locator.first.click(timeout=self.config.timeout_ms)
                            new_page = pinfo.value
                            new_page.wait_for_load_state("domcontentloaded", timeout=self.config.timeout_ms)
                            self.page_stack.append(self.page)
                            self.page = new_page
                        else:
                            locator.first.click(timeout=self.config.timeout_ms)
                    else:
                        locator.first.wait_for(state="visible", timeout=self.config.timeout_ms)
                elif action == "type":
                    if not locator:
                        raise RuntimeError("Locator not resolved")
                    if mode != "dry-run":
                        locator.first.fill(str(value), timeout=self.config.timeout_ms)
                    else:
                        locator.first.wait_for(state="visible", timeout=self.config.timeout_ms)
                elif action == "select":
                    if not locator:
                        raise RuntimeError("Locator not resolved")
                    if mode != "dry-run":
                        try:
                            locator.first.select_option(label=str(value))
                        except Exception:
                            locator.first.click(timeout=self.config.timeout_ms)
                            locator.first.fill(str(value))
                            locator.first.press("Enter")
                    else:
                        locator.first.wait_for(state="visible", timeout=self.config.timeout_ms)
                elif action == "wait":
                    if locator:
                        locator.first.wait_for(state=wait_state, timeout=self.config.timeout_ms)
                    else:
                        time.sleep(float(value) if value else 1.0)
                elif action == "assert":
                    if locator:
                        locator.first.wait_for(state="visible", timeout=self.config.timeout_ms)
                    elif value:
                        self.page.get_by_text(str(value), exact=False).first.wait_for(
                            state="visible", timeout=self.config.timeout_ms
                        )
                elif action == "download":
                    if not locator:
                        raise RuntimeError("Locator not resolved")
                    if mode != "dry-run":
                        with self.page.expect_download(timeout=self.config.timeout_ms) as dl_info:
                            locator.first.click(timeout=self.config.timeout_ms)
                        download = dl_info.value
                        ensure_dir(self.config.download_dir)
                        target = self.config.download_dir / (download.suggested_filename or "download.bin")
                        download.save_as(str(target))
                        self.logger.info("Download saved: %s", target)
                    else:
                        locator.first.wait_for(state="visible", timeout=self.config.timeout_ms)
                elif action == "upload":
                    file_path = Path(str(value)) if value else None
                    if not file_path or not file_path.exists():
                        raise RuntimeError(f"File not found: {file_path}")
                    if mode != "dry-run":
                        if locator:
                            locator.first.set_input_files(str(file_path))
                        else:
                            self.page.set_input_files("input[type='file']", str(file_path))
                    else:
                        if locator:
                            locator.first.wait_for(state="visible", timeout=self.config.timeout_ms)
                elif action == "select_row":
                    if not locator:
                        raise RuntimeError("Row not resolved")
                    if mode != "dry-run":
                        locator.scroll_into_view_if_needed()
                        locator.click(timeout=self.config.timeout_ms)
                    else:
                        locator.wait_for(state="visible", timeout=self.config.timeout_ms)
                elif action == "filter_grid":
                    if not target_hint:
                        raise RuntimeError("Column name not provided")
                    grid_hint = expand_value(locator_hint, self.variables)
                    grid_candidates = split_candidates(grid_hint) if grid_hint else []
                    input_loc = None
                    for grid_sel in grid_candidates:
                        input_loc = self.find_grid_filter_input(grid_sel, target_hint)
                        if input_loc and input_loc.count() > 0:
                            break
                    if not input_loc:
                        raise RuntimeError("Grid filter input not resolved")
                    if mode != "dry-run":
                        input_loc.fill(str(value), timeout=self.config.timeout_ms)
                        try:
                            input_loc.press("Enter")
                        except Exception:
                            pass
                    else:
                        input_loc.wait_for(state="visible", timeout=self.config.timeout_ms)
                elif action == "click_row_icon":
                    icon = self.resolve_row_icon(expand_value(locator_hint, self.variables))
                    if not icon:
                        raise RuntimeError("Row icon not resolved")
                    if mode != "dry-run":
                        icon.scroll_into_view_if_needed()
                        icon.click(timeout=self.config.timeout_ms)
                    else:
                        icon.wait_for(state="visible", timeout=self.config.timeout_ms)
                elif action == "close_page":
                    if self.page_stack:
                        try:
                            self.page.close()
                        except Exception:
                            pass
                        self.page = self.page_stack.pop()
                else:
                    raise RuntimeError(f"Unsupported action: {action}")

                self.wait_for_idle(self.config.timeout_ms)
                if evidence:
                    self.take_evidence(step_id, target_hint or action)
                return
            except Exception as exc:
                if attempt >= retries:
                    self.save_failure_artifacts(step_id, target_hint or action)
                    if optional:
                        self.logger.warning("Optional step failed: %s", exc)
                        return
                    raise
                time.sleep(0.8)


def run_steps(steps: list[dict[str, Any]], runner: StepRunner, mode: str, processes: list[str]) -> None:
    global_steps = [s for s in steps if s.get("scope") != "per_process"]
    process_steps = [s for s in steps if s.get("scope") == "per_process"]

    for step in global_steps:
        runner.run_step(step, mode=mode)
        if mode == "debug":
            input("Press Enter to continue...")

    if not process_steps:
        return

    for idx, processo in enumerate(processes, start=1):
        runner.logger.info("Processo %s/%s: %s", idx, len(processes), processo)
        runner.set_process(processo)
        for step in process_steps:
            runner.run_step(step, mode=mode)
            if mode == "debug":
                input("Press Enter to continue...")


def build_process_list(config) -> list[str]:
    if config.single_processo:
        return [config.single_processo]

    if config.process_list:
        processes = config.process_list
    elif config.process_all:
        latest = find_latest_excel(config.download_dir)
        if not latest:
            return []
        processes = list(extract_processes_from_excel(latest))
    else:
        processes = []

    if config.max_processes and config.max_processes > 0:
        processes = processes[: config.max_processes]
    return processes


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(description="Run the video-driven RPA steps.")
    parser.add_argument("--mode", choices=["run", "debug", "dry-run"], help="Execution mode.")
    parser.add_argument("--steps", help="Path to steps.yaml.")
    parser.add_argument("--processo", help="Single process number to run.")
    parser.add_argument("--processos", help="Comma or semicolon-separated list of processos.")
    parser.add_argument("--process-all", action="store_true", help="Use all processos from exported Excel.")
    parser.add_argument("--max-processos", type=int, help="Max processos to run.")
    args = parser.parse_args()

    config = load_config()
    if args.mode:
        config.mode = args.mode
    if args.steps:
        config.steps_path = Path(args.steps)
    if args.processo:
        config.single_processo = args.processo
    if args.processos:
        config.process_list = [s.strip() for s in args.processos.replace(";", ",").split(",") if s.strip()]
    if args.process_all:
        config.process_all = True
    if args.max_processos is not None:
        config.max_processes = args.max_processos
    logger = init_logger(config.log_dir)
    if not config.etcm_user or not config.etcm_pass:
        logger.warning("Missing ETCM_USER/ETCM_PASS (login may require manual fill).")

    steps = read_yaml_steps(config.steps_path)
    mode = config.mode
    if mode not in {"run", "debug", "dry-run"}:
        mode = "run"

    ensure_dir(config.artifacts_dir)
    ensure_dir(config.download_dir)
    ensure_dir(config.evidence_dir)
    ensure_dir(config.html_dir)

    processes = build_process_list(config)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=config.headless, slow_mo=config.slowmo_ms, channel="chrome")
        context_kwargs = {"accept_downloads": True, "viewport": {"width": 1600, "height": 900}}
        if config.use_storage_state and config.storage_state_path.exists():
            context_kwargs["storage_state"] = str(config.storage_state_path)
        context = browser.new_context(**context_kwargs)
        page = context.new_page()
        page.set_default_timeout(config.timeout_ms)

        runner = StepRunner(page, context, logger, config)

        try:
            run_steps(steps, runner, mode=mode, processes=processes)
        finally:
            if config.use_storage_state:
                try:
                    context.storage_state(path=str(config.storage_state_path))
                except Exception:
                    pass
            context.close()
            browser.close()


if __name__ == "__main__":
    main()
